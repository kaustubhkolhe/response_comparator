import difflib
import textdistance
from fuzzywuzzy import fuzz
import jellyfish
import openpyxl
import nltk


def print_hyperlinks_and_values(excel_file):
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except FileNotFoundError:
        print(f"Error: File '{excel_file}' not found.")
        return

    question_column = 1  # Assuming questions are in the first column
    actual_column = 2  # Assuming actual answers are in the second column
    expected_column = 3  # Assuming expected answers are in the third column

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Iterate through all rows in the sheet, starting from the second row
        for row in sheet.iter_rows(min_row=2):

            # Access the cells in the specified columns
            question = row[question_column - 1].value.lower().strip()
            actual_cell = row[actual_column - 1]
            expected_cell = row[expected_column - 1]

            # Print actual value and hyperlink if present
            if actual_cell.hyperlink:
                actual_answer = str(actual_cell.hyperlink.target)
            else:
                actual_answer = actual_cell.value

            # Print expected value and hyperlink if present
            if expected_cell.hyperlink:
                expected_answer = str(expected_cell.hyperlink.target)
            else:
                expected_answer = expected_cell.value

            # Difflib comparison
            seq_matcher = difflib.SequenceMatcher(None, actual_answer.lower(), expected_answer.lower())
            difflib_score = seq_matcher.ratio()

            # TextDistance comparison
            textdistance_score = textdistance.levenshtein.normalized_similarity(actual_answer.lower(),
                                                                                expected_answer.lower())

            # FuzzyWuzzy comparison
            fuzzywuzzy_score = fuzz.ratio(actual_answer.lower(), expected_answer.lower()) / 100

            # Jellyfish comparison
            jellyfish_score = jellyfish.jaro_winkler_similarity(actual_answer.lower(), expected_answer.lower())

            # NLTK comparison
            nltk_score = 1.0 - nltk.edit_distance(actual_answer.lower(), expected_answer.lower()) / max(
                len(actual_answer),
                len(expected_answer))

            print(f"Question: {question}")
            print(f"Actual Answer: {actual_answer}")
            print(f"Expected Answer: {expected_answer}")
            print(f"Difflib Score: {difflib_score:.2f}")
            print(f"TextDistance Score: {textdistance_score:.2f}")
            print(f"FuzzyWuzzy Score: {fuzzywuzzy_score:.2f}")
            print(f"Jellyfish Score: {jellyfish_score:.2f}")
            print(f"NLTK Score: {nltk_score:.2f}\n")
            print(
                "_____________________________________________________________________________________________________")


excel_file = 'Q&A_data.xlsx'
print_hyperlinks_and_values(excel_file)
