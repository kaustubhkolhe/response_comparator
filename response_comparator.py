import difflib
import textdistance
from fuzzywuzzy import fuzz
import jellyfish
import openpyxl
import nltk
from openpyxl.styles import Font
import re

def preprocess_text(text):
    # Define a regex pattern to match numbered patterns and various types of separators
    separator_pattern = r'^\s*(?:\d+[\.\)]|[\u2022\u2023\u25E6\-•]\s*)'
    # Replace separators with a single space
    text = re.sub(separator_pattern, '', text, flags=re.MULTILINE)
    # Remove extra spaces and line breaks
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def print_hyperlinks_and_values(excel_file):
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except FileNotFoundError:
        print(f"Error: File '{excel_file}' not found.")
        return

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Add headers for the scores
        sheet.cell(row=1, column=4).value = "Difflib Score"
        sheet.cell(row=1, column=5).value = "TextDistance Score"
        sheet.cell(row=1, column=6).value = "FuzzyWuzzy Score"
        sheet.cell(row=1, column=7).value = "Jellyfish Score"
        sheet.cell(row=1, column=8).value = "NLTK Score"

        # Bold the first row
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Iterate through all rows in the sheet, starting from the second row
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):

            # Access the cells in the specified columns
            actual_cell = row[1]
            expected_cell = row[2]

            # Print actual value and hyperlink if present
            if actual_cell.hyperlink:
                actual_answer = str(actual_cell.hyperlink.target)
            else:
                actual_answer = actual_cell.value

            # Print expected value and hyperlink if present
            if expected_cell.hyperlink:
                expected_answer = str(expected_cell.hyperlink.target)
            else:
                expected_answer = expected_cell.value

            # Preprocess text
            actual_answer = preprocess_text(actual_answer.lower())
            expected_answer = preprocess_text(expected_answer.lower())

            # Difflib comparison
            seq_matcher = difflib.SequenceMatcher(None, actual_answer, expected_answer)
            difflib_score = seq_matcher.ratio()

            # TextDistance comparison
            textdistance_score = textdistance.levenshtein.normalized_similarity(actual_answer, expected_answer)

            # FuzzyWuzzy comparison
            fuzzywuzzy_score = fuzz.ratio(actual_answer, expected_answer) / 100

            # Jellyfish comparison
            jellyfish_score = jellyfish.jaro_winkler_similarity(actual_answer, expected_answer)

            # NLTK comparison
            nltk_score = 1.0 - nltk.edit_distance(actual_answer, expected_answer) / max(
                len(actual_answer),
                len(expected_answer))

            # Write scores to the Excel sheet
            sheet.cell(row=idx, column=4).value = difflib_score
            sheet.cell(row=idx, column=5).value = textdistance_score
            sheet.cell(row=idx, column=6).value = fuzzywuzzy_score
            sheet.cell(row=idx, column=7).value = jellyfish_score
            sheet.cell(row=idx, column=8).value = nltk_score

    # Save the modified Excel file
    wb.save(excel_file)

excel_file = 'Q&A_data.xlsx'
print_hyperlinks_and_values(excel_file)
